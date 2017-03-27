# -*- coding: utf-8 -*-
from functools import wraps
from flask import render_template, flash, redirect, url_for, request, abort, jsonify
from sqlalchemy import desc, asc, or_
from flask_login import login_user, logout_user, current_user, login_required
from datetime import datetime, timedelta
from app import app, models, db, forms
from flask_mobility.decorators import mobile_template
import os
from werkzeug.utils import secure_filename
from config import ALLOWED_EXTENSIONS
from flask import send_from_directory
import json


@app.context_processor
def utility_processor():
    def getthumb(path):
        import os
        filename, file_extension = os.path.splitext(path)
        return filename + '_thumb' + file_extension

    return dict(getthumb=getthumb)


def webLog(func):
    @wraps(func)
    def newFunc(*args, **kwargs):
        if request.headers.getlist("X-Forwarded-For"):
            ip = request.headers.getlist("X-Forwarded-For")[0]
        else:
            ip = request.remote_addr
        db.session.add(models.Logger(url=request.url, remote_addr=ip, method=request.method,
                                     user_agent=request.user_agent, datetime=datetime.now()))
        db.session.commit()
        return func(*args, **kwargs)

    return newFunc


@app.route('/')
@app.route('/index')
@webLog
def index():
    from urllib.parse import urlparse
    newsslider = models.News_Slider.query.order_by(desc(models.News_Slider.timestamp)).limit(3).all()
    video = models.Video_slider.query.order_by(desc(models.Video_slider.date_added)).filter_by(active=True).limit(
        12).all()
    params = []
    for url in video:
        parsedlink = urlparse(url.url)
        if parsedlink.netloc == 'youtu.be' or parsedlink.netloc == 'www.youtu.be':
            params.append(parsedlink.path.replace('/', ''))
        else:
            params.append(parsedlink.query.replace('v=', ''))
    posts = models.Post.query.order_by(desc(models.Post.timestamp)).limit(9).all()
    return render_template('index.html', posts=posts, video=params, newsslider=newsslider)


@app.route('/login', methods=['POST'])
@webLog
def login():
    username = request.form['username']
    password = request.form['password']
    registered_user = models.User.query.filter_by(login=username, password=password).first()
    if registered_user is None:
        flash('Username or Password is invalid', 'error')
        return redirect(url_for('index'))
    login_user(registered_user)
    flash('Logged in successfully')
    return redirect(request.args.get('next') or url_for('index'))


@app.route('/logout/')
@webLog
def logout_view():
    logout_user()
    return redirect(url_for('index'))


@app.route('/registration', methods=['POST'])
@webLog
def invite_reg():
    person = models.Person.query.filter_by(invite=request.form['invite']).first()
    if person:
        return render_template('reg_user.html', user=person)
    else:
        return abort(404)


@app.route('/add_user', methods=['POST'])
@webLog
def add_user():
    person = models.Person.query.filter_by(id=request.form['id']).first()
    login = request.form['login']
    email = request.form['email']
    password = request.form['password']
    db.session.add(models.User(login=login, email=email, password=password, person_id=person.id))
    db.session.commit()
    registered_user = models.User.query.filter_by(login=login, password=password).first()
    if registered_user is None:
        flash('Username or Password is invalid', 'error')
        return redirect(url_for('invite_reg'))
    login_user(registered_user)
    person.invite = None
    db.session.commit()
    flash('Logged in successfully')
    return redirect(url_for('profile'))


@app.route('/profile')
@webLog
def profile():
    paypay = {'Sep': 0, 'Oct': 0, 'Nov': 0, 'Dec': 0, 'Jan': 0, 'Feb': 0, 'Mar': 0, 'Apr': 0, 'May': 0, 'Jun': 0,
              'Jul': 0, 'Aug': 0}
    person = models.Person.query.filter_by(id=current_user.person_id).first()
    payment = models.Payment.query.filter_by(person=person.id).all()
    for pay in payment:
        if pay.date.month == 9:
            paypay['Sep'] = 1
            paypay['Sep_price'] = pay.price
        elif pay.date.month == 10:
            paypay['Oct'] = 1
            paypay['Oct_price'] = pay.price
        elif pay.date.month == 11:
            paypay['Nov'] = 1
            paypay['Nov_price'] = pay.price
        elif pay.date.month == 12:
            paypay['Dec'] = 1
            paypay['Dec_price'] = pay.price
        elif pay.date.month == 1:
            paypay['Jan'] = 1
            paypay['Jan_price'] = pay.price
        elif pay.date.month == 2:
            paypay['Feb'] = 1
            paypay['Feb_price'] = pay.price
        elif pay.date.month == 3:
            paypay['Mar'] = 1
            paypay['Mar_price'] = pay.price
        elif pay.date.month == 4:
            paypay['Apr'] = 1
            paypay['Apr_price'] = pay.price
        elif pay.date.month == 5:
            paypay['May'] = 1
            paypay['May_price'] = pay.price
        elif pay.date.month == 6:
            paypay['Jun'] = 1
            paypay['Jun_price'] = pay.price
        elif pay.date.month == 7:
            paypay['Jul'] = 1
            paypay['Jul_price'] = pay.price
        elif pay.date.month == 8:
            paypay['Aug'] = 1
            paypay['Aug_price'] = pay.price

    work = models.Work.query.filter_by(person=person.id).all()
    hours = timedelta()
    for w in work:
        hours += w.end - w.start

    violations = models.Violation.query.filter_by(person=person.id).all()

    wperson = models.Person.query.filter_by(id=current_user.person_id).first()
    if wperson:
        hostel = db.session.query(models.Hostel).filter(models.Room.id == wperson.room).filter(
            models.Hostel.id == models.Room.hostel_id).first().id
        washing = models.Washing.query.filter_by(hostel=hostel).all()
    else:
        hostel = None
        washing = models.Washing.query.all()

    data = []

    if washing:
        for wash in washing:
            data.append({'id': wash.id, 'name': wash.person, 'location': wash.hostel,
                         'startDate': wash.start.strftime('%m/%d/%Y'), 'endDate': wash.end.strftime('%m/%d/%Y')})

    return render_template('profile.html', person=person, payment=paypay, hours=hours, work=work, violations=violations,
                           data=data, wperson=wperson, hostel=hostel)


@login_required
@app.route('/washing', methods=['GET', 'POST'])
@webLog
def washing():
    hostel_id = models.Room.query.filter_by(
        id=models.Person.query.filter_by(id=current_user.person_id).first().room).first().hostel_id
    days_list = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    start = datetime(year=datetime.now().year, month=datetime.now().month, day=datetime.now().day, hour=0, minute=0,
                     second=0, microsecond=1)
    end = start + timedelta(days=6 - start.weekday(), hours=23, minutes=59, seconds=59, microseconds=59)
    days = start.weekday()
    washing = models.Washing.query.filter(models.Washing.start > str(start)).filter(
        models.Washing.end < str(end)).filter(models.Washing.hostel == hostel_id).all()
    washing_ = {'Monday': [], 'Tuesday': [], 'Wednesday': [], 'Thursday': [], 'Friday': [], 'Saturday': [],
                'Sunday': []}
    for wash in washing:
        washing_[wash.start.strftime("%A")].append([wash.start, wash.end, models.Person.query.filter_by(
            id=wash.person).first().last_name + ' ' + models.Person.query.filter_by(id=wash.person).first().first_name])
    return render_template('washing.html', washing=washing_, days=days, days_list=days_list)


@app.route('/post/<id>')
@webLog
def detail_view_post(id):
    post = models.Post.query.filter_by(id=id).first()
    if post:
        return render_template('post.html', post=post)
    else:
        abort(404)


@app.route('/hostels', methods=['GET', 'POST'])
@webLog
def rooms():
    form = forms.SearchForm(request.form)
    form2 = forms.SearchForm2(request.form)
    if request.method == 'GET':
        hostels = models.Hostel.query.all()
        return render_template('hostels.html', hostels=hostels, form=form, form2=form2)
    elif request.method == 'POST':
        if request.form['button'] == 'secondS':
            return redirect(url_for('room_detail', hostel=form2.hostel_id.data, room=form2.room_id.data))
        elif request.form['button'] == 'firstS':
            kwargs = {form.radio.data: form.value.data}
            search_result = models.Person.query.filter_by(**kwargs).all()
            return render_template('search_person.html', result=search_result)


@app.route('/hostels/<hostel>')
@webLog
def hostel_detail(hostel):
    hostel_number = hostel
    hostel = models.Hostel.query.filter_by(number=hostel).first().id
    if int(hostel_number) == 2:
        _rooms = models.Room.query.filter_by(hostel_id=hostel).all()
        floors = models.Room.query.order_by(models.Room.floor).filter_by(hostel_id=hostel).group_by(
            models.Room.floor).all()
        places = {}
        for free in models.Room_free.query.all():
            places.update({str(free.room_id): free.places})

        return render_template('hostel_view.html', hostel_number=hostel_number, rooms=_rooms, floors=floors,
                               room_free=places)
    else:
        blocks = models.Block.query.filter_by(hostel_id=hostel)
        floors = models.Block.query.order_by(models.Block.floor).filter_by(hostel_id=hostel).group_by(
            models.Block.floor).all()
    return render_template('blocks.html', hostel_number=hostel_number, blocks=blocks, floors=floors)


@app.route('/hostel/<hostel>/<block>')
@webLog
def block_view(hostel, block):
    hostel_number = hostel
    block_number = block
    hostel_id = models.Hostel.query.filter_by(number=hostel).first().id
    block_id = models.Block.query.filter_by(number=block_number).first().id
    rooms = models.Room.query.filter_by(hostel_id=hostel_id, block_id=block_id).all()
    query = [x.id for x in rooms]
    places = {}
    for free in db.session.query(models.Room_free).filter(models.Room_free.room_id.in_(query)).all():
        places.update({str(free.room_id): free.places})

    return render_template('block_view.html', hostel_number=hostel_number, rooms=rooms, room_free=places)


@app.route('/hostels/<hostel>/free')
@webLog
def hostel_detail_free(hostel):
    places = models.Room.query.filter_by(hostel_id=hostel).all()
    places_free = []
    for place in places:
        person_buffer = models.Person.query.filter_by(hostel_id=hostel, room_id=place.room_number).all()
        room_buffer = models.Room.query.filter_by(hostel_id=hostel, room_number=place.room_number).first()
        if room_buffer.numbers_of_person is None:
            room_buffer.numbers_of_person = 0
        if int(room_buffer.numbers_of_person) - len(person_buffer) > 0:
            places_free.append(room_buffer)
    floors = models.Room.query.order_by(models.Room.floor).filter_by(hostel_id=hostel).group_by(models.Room.floor).all()

    return render_template('free.html', hostel=hostel, places_free=places_free, floors=floors)


@app.route('/rooms/<hostel>/<room>')
@webLog
def room_detail(hostel, room):
    hostel_number = hostel
    hostel = models.Hostel.query.filter_by(number=hostel).first().id
    room = models.Room.query.filter_by(hostel_id=hostel, room_number=room).first().id
    persons = models.Person.query.filter_by(room=room).all()
    places = models.Room.query.filter_by(hostel_id=hostel, id=room).first()
    if places is not None:
        return render_template('room_view.html', hostel_number=hostel_number, persons=persons, places=places)
    else:
        abort(404)


@app.route('/register', methods=['GET', 'POST'])
@webLog
def register():
    form = forms.RegistrationForm(request.form)
    if request.method == 'POST' and form.validate():
        main_info = [form.room_type.data, form.reason.data, form.first_name.data, form.last_name.data,
                     form.middle_name.data, form.birthday.data, form.department.data, form.specialty.data,
                     form.group.data, form.work.data, form.s_passport.data, form.n_passport.data, form.d_passport.data,
                     form.k_passport.data, form.phone_number.data, form.lived_hostel.data, form.lived_room.data,
                     form.form_of_education.data]
        main_info = [None if x == '' else x for x in main_info]
        main_info = models.Register_main(*main_info)
        db.session.add(main_info)
        db.session.commit()

        if form.family_radio.data == 'y':
            main_family = [main_info.id, form.husband_wife.data, form.husband_wife_work.data,
                           form.husband_wife_birthday.data,
                           form.husband_wife_s_passport.data, form.husband_wife_n_passport.data,
                           form.husband_wife_d_passport.data, form.husband_wife_k_passport.data,
                           form.husband_wife_form_of_education.data, form.husband_wife_lived.data,
                           form.husband_wife_lived_hostel.data, form.husband_wife_lived_room.data, form.childrens.data,
                           form.children_live.data]
            main_family = [None if x == '' else x for x in main_family]
            db.session.add(models.Register_family(*main_family))
            db.session.commit()

        elif form.family_radio.data == 'n':
            main_student = [main_info.id, form.father.data, form.father_work.data, form.mother.data,
                            form.mother_work.data, form.brothers_sisters.data, form.parents_street.data,
                            form.parents_home.data, form.parents_apartment.data, form.parents_settlement.data,
                            form.parents_district.data, form.parents_region.data, form.parents_index.data,
                            form.parents_landline_phone.data, form.parents_mobile_phone.data]
            main_student = [None if x == '' else x for x in main_student]
            db.session.add(models.Register_student(*main_student))
            db.session.commit()
        flash('Thanks for registering')
        return redirect(url_for('index'))
    return render_template('register.html', form=form)


@app.route('/all_register')
@webLog
def all_register():
    family = db.session.query(models.Register_main, models.Register_family) \
        .filter(models.Register_family.register_id == models.Register_main.id).all()
    student = db.session.query(models.Register_main, models.Register_student) \
        .filter(models.Register_student.register_id == models.Register_main.id).all()
    return render_template('all_register.html', family=family, student=student)


@app.route('/plot')
@webLog
def plot():
    temp = db.session.query(models.Temperature).order_by(asc(models.Temperature.date)).all()
    buffer = {}
    min = temp[0].date.strftime('%m/%d/%Y')
    max = temp[-1].date.strftime('%m/%d/%Y')
    for t in temp:
        if buffer.get(str(t.date)):
            buffer.get(str(t.date)).update({t.hostel_id: t.temperature})
        else:
            buffer.update({str(t.date): {t.hostel_id: t.temperature}})

    return render_template('plot.html', min=min, max=max)


@app.route('/temp_xlsx', methods=['GET', 'POST'])
@webLog
def temp_xlsx():
    if request.method == 'GET':
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        temp = db.session.query(models.Temperature).order_by(asc(models.Temperature.date)).all()
        buffer = {}
        for t in temp:
            if buffer.get(t.date):
                buffer.get(t.date).update({t.hostel_id: t.temperature})
            else:
                buffer.update({t.date: {t.hostel_id: t.temperature}})

        ws.append(['Дата', 'Общежитие 2', 'Общежите 3', 'Общежитие 4'])
        for i in sorted(buffer):
            ws.append([i, buffer.get(i).get(2), buffer.get(i).get(3), buffer.get(i).get(4)])

        wb.save('app/temp.xlsx')
        return send_from_directory('', 'temp.xlsx')

    else:
        f = request.files['file']
        if f.filename.split('.')[-1] == 'xlsx':
            f.save(f.filename)
            from openpyxl import load_workbook
            wb = load_workbook(f.filename)
            ws = wb.worksheets[0]
            if request.form['row'] and type(request.form['row'] == int):
                end = request.form['row']
            else:
                end = ws.max_row + 1
            for i in range(2, end):
                h2 = models.Temperature.query.filter_by(date=ws['A%s' % i].value.date(), hostel_id=2).first()
                h3 = models.Temperature.query.filter_by(date=ws['A%s' % i].value.date(), hostel_id=3).first()
                h4 = models.Temperature.query.filter_by(date=ws['A%s' % i].value.date(), hostel_id=4).first()
                if h2:
                    h2.temperature = ws['B%s' % i].value
                else:
                    db.session.add(
                        models.Temperature(date=ws['A%s' % i].value, temperature=ws['B%s' % i].value, hostel_id=2))
                if h3:
                    h3.temperature = ws['C%s' % i].value
                else:
                    db.session.add(
                        models.Temperature(date=ws['A%s' % i].value, temperature=ws['C%s' % i].value, hostel_id=3))
                if h4:
                    h4.temperature = ws['D%s' % i].value
                else:
                    db.session.add(
                        models.Temperature(date=ws['A%s' % i].value, temperature=ws['D%s' % i].value, hostel_id=4))
                db.session.commit()
        else:
            print('Error file extension')
        return redirect(url_for('plot'))


@app.route('/plot_range', methods=['POST'])
@webLog
def plot_range():
    from collections import OrderedDict
    start = request.form['start'].split('-')
    end = request.form['end'].split('-')
    q = models.Temperature.query.filter(
        models.Temperature.date >= datetime.date(datetime(year=int(start[0]), month=int(start[1]), day=int(start[2])))) \
        .filter(
        models.Temperature.date <= datetime.date(datetime(year=int(end[0]), month=int(end[1]), day=int(end[2])))) \
        .order_by(asc(models.Temperature.date)).all()
    buffer = OrderedDict()
    for t in q:
        if buffer.get(str(t.date)):
            buffer.get(str(t.date)).update({t.hostel_id: t.temperature})
        else:
            buffer.update({str(t.date): {t.hostel_id: t.temperature}})
    return json.dumps(buffer)


@app.route('/person/<id>')
@webLog
def person(id):
    person_ = models.Person.query.filter_by(id=id).first()
    return render_template('person.html', person=person_)


@app.route('/stat', methods=['GET', 'POST'])
@webLog
def stat():
    if request.method == 'POST':
        last = models.Statistics.query.order_by(desc(models.Statistics.date)).first()
        if last.date == datetime.date(datetime.now()):
            flash('Сегодя уже было обновление!')
            return redirect('stat')
        else:
            free_1 = 0
            free_2 = 0
            free_3 = 0
            free_4 = 0
            places_all = models.Room.query.all()
            places2 = models.Room.query.filter_by(hostel_id=2).all()
            places3 = models.Room.query.filter_by(hostel_id=3).all()
            places4 = models.Room.query.filter_by(hostel_id=4).all()
            for places in places_all:
                person_buffer = models.Person.query.filter_by(hostel_id=places.hostel_id,
                                                              room_id=places.room_number).all()
                room_buffer = models.Room.query.filter_by(hostel_id=places.hostel_id,
                                                          room_number=places.room_number).first()
                if room_buffer.numbers_of_person is None:
                    room_buffer.numbers_of_person = 0
                if int(room_buffer.numbers_of_person) - len(person_buffer) == 1:
                    free_1 += 1
                elif int(room_buffer.numbers_of_person) - len(person_buffer) == 2:
                    free_2 += 1
                elif int(room_buffer.numbers_of_person) - len(person_buffer) == 3:
                    free_3 += 1
                elif int(room_buffer.numbers_of_person) - len(person_buffer) == 4:
                    free_4 += 1
            stats = models.Statistics(datetime.date(datetime.now()), len(places_all), len(places2), len(places3),
                                      len(places3), free_1, free_2, free_3, free_4)
            db.session.add(stats)
            db.session.commit()
            return redirect('stat')
    elif request.method == 'GET':
        stats = models.Statistics.query.order_by(desc(models.Statistics.date)).first()

        departments = [
            {'department': 'ФІФ', 'len': len(models.Person.query.filter_by(department='ФІФ').all())},
            {'department': 'ФБГЕ', 'len': len(models.Person.query.filter_by(department='ФБГЕ').all())},
            {'department': 'ФДПО', 'len': len(models.Person.query.filter_by(department='ФДПО').all())},
            {'department': 'ФЕМ', 'len': len(models.Person.query.filter_by(department='ФЕМ').all())},
            {'department': 'ФКМ', 'len': len(models.Person.query.filter_by(department='ФКМ').all())},
            {'department': 'ФПІС', 'len': len(models.Person.query.filter_by(department='ФПІС').all())},
            {'department': 'ФПЗЛТ', 'len': len(models.Person.query.filter_by(department='ФПЗЛТ').all())},
            {'department': 'ФТСО', 'len': len(models.Person.query.filter_by(department='ФТСО').all())},
            {'department': 'ФФВС', 'len': len(models.Person.query.filter_by(department='ФФВС').all())},
            {'department': 'ФФЖ', 'len': len(models.Person.query.filter_by(department='ФФЖ').all())},
            {'department': 'ФФМІ', 'len': len(models.Person.query.filter_by(department='ФФМІ').all())},
            {'department': 'ЮФ', 'len': len(models.Person.query.filter_by(department='ЮФ').all())},
            {'department': 'переклад', 'len': len(models.Person.query.filter_by(department='переклад').all())},
        ]

        departments_2 = [
            {'department': 'ФІФ', 'len': len(models.Person.query.filter_by(department='ФІФ', hostel_id=1).all())},
            {'department': 'ФБГЕ', 'len': len(models.Person.query.filter_by(department='ФБГЕ', hostel_id=1).all())},
            {'department': 'ФДПО', 'len': len(models.Person.query.filter_by(department='ФДПО', hostel_id=1).all())},
            {'department': 'ФЕМ', 'len': len(models.Person.query.filter_by(department='ФЕМ', hostel_id=1).all())},
            {'department': 'ФКМ', 'len': len(models.Person.query.filter_by(department='ФКМ', hostel_id=1).all())},
            {'department': 'ФПІС', 'len': len(models.Person.query.filter_by(department='ФПІС', hostel_id=1).all())},
            {'department': 'ФПЗЛТ', 'len': len(models.Person.query.filter_by(department='ФПЗЛТ', hostel_id=1).all())},
            {'department': 'ФТСО', 'len': len(models.Person.query.filter_by(department='ФТСО', hostel_id=1).all())},
            {'department': 'ФФВС', 'len': len(models.Person.query.filter_by(department='ФФВС', hostel_id=1).all())},
            {'department': 'ФФЖ', 'len': len(models.Person.query.filter_by(department='ФФЖ', hostel_id=1).all())},
            {'department': 'ФФМІ', 'len': len(models.Person.query.filter_by(department='ФФМІ', hostel_id=1).all())},
            {'department': 'ЮФ', 'len': len(models.Person.query.filter_by(department='ЮФ', hostel_id=1).all())},
            {'department': 'переклад',
             'len': len(models.Person.query.filter_by(department='переклад', hostel_id=1).all())},
        ]

        departments_3 = [
            {'department': 'ФІФ', 'len': len(models.Person.query.filter_by(department='ФІФ', hostel_id=2).all())},
            {'department': 'ФБГЕ', 'len': len(models.Person.query.filter_by(department='ФБГЕ', hostel_id=2).all())},
            {'department': 'ФДПО', 'len': len(models.Person.query.filter_by(department='ФДПО', hostel_id=2).all())},
            {'department': 'ФЕМ', 'len': len(models.Person.query.filter_by(department='ФЕМ', hostel_id=2).all())},
            {'department': 'ФКМ', 'len': len(models.Person.query.filter_by(department='ФКМ', hostel_id=2).all())},
            {'department': 'ФПІС', 'len': len(models.Person.query.filter_by(department='ФПІС', hostel_id=2).all())},
            {'department': 'ФПЗЛТ', 'len': len(models.Person.query.filter_by(department='ФПЗЛТ', hostel_id=2).all())},
            {'department': 'ФТСО', 'len': len(models.Person.query.filter_by(department='ФТСО', hostel_id=2).all())},
            {'department': 'ФФВС', 'len': len(models.Person.query.filter_by(department='ФФВС', hostel_id=2).all())},
            {'department': 'ФФЖ', 'len': len(models.Person.query.filter_by(department='ФФЖ', hostel_id=2).all())},
            {'department': 'ФФМІ', 'len': len(models.Person.query.filter_by(department='ФФМІ', hostel_id=2).all())},
            {'department': 'ЮФ', 'len': len(models.Person.query.filter_by(department='ЮФ', hostel_id=2).all())},
            {'department': 'переклад',
             'len': len(models.Person.query.filter_by(department='переклад', hostel_id=2).all())},
        ]

        departments_4 = [
            {'department': 'ФІФ', 'len': len(models.Person.query.filter_by(department='ФІФ', hostel_id=3).all())},
            {'department': 'ФБГЕ', 'len': len(models.Person.query.filter_by(department='ФБГЕ', hostel_id=3).all())},
            {'department': 'ФДПО', 'len': len(models.Person.query.filter_by(department='ФДПО', hostel_id=3).all())},
            {'department': 'ФЕМ', 'len': len(models.Person.query.filter_by(department='ФЕМ', hostel_id=3).all())},
            {'department': 'ФКМ', 'len': len(models.Person.query.filter_by(department='ФКМ', hostel_id=3).all())},
            {'department': 'ФПІС', 'len': len(models.Person.query.filter_by(department='ФПІС', hostel_id=3).all())},
            {'department': 'ФПЗЛТ', 'len': len(models.Person.query.filter_by(department='ФПЗЛТ', hostel_id=3).all())},
            {'department': 'ФТСО', 'len': len(models.Person.query.filter_by(department='ФТСО', hostel_id=3).all())},
            {'department': 'ФФВС', 'len': len(models.Person.query.filter_by(department='ФФВС', hostel_id=3).all())},
            {'department': 'ФФЖ', 'len': len(models.Person.query.filter_by(department='ФФЖ', hostel_id=3).all())},
            {'department': 'ФФМІ', 'len': len(models.Person.query.filter_by(department='ФФМІ', hostel_id=3).all())},
            {'department': 'ЮФ', 'len': len(models.Person.query.filter_by(department='ЮФ', hostel_id=3).all())},
            {'department': 'переклад',
             'len': len(models.Person.query.filter_by(department='переклад', hostel_id=3).all())},
        ]

        rooms = [{'hostel': 2, 'len': len(models.Room.query.filter_by(hostel_id=1).all())},
                 {'hostel': 3, 'len': len(models.Room.query.filter_by(hostel_id=2).all())},
                 {'hostel': 4, 'len': len(models.Room.query.filter_by(hostel_id=3).all())}]

        courses = [{'course': '1 курс', 'len': 0},
                   {'course': '2 курс', 'len': 0},
                   {'course': '3 курс', 'len': 0},
                   {'course': '4 курс', 'len': 0},
                   {'course': '5 курс', 'len': 0},
                   {'course': '6 курс', 'len': 0}]

        query = models.Person.query.all()
        for c in query:
            if c.group:
                if str(c.group)[:1] == '1':
                    courses[0]['len'] += 1
                elif str(c.group)[:1] == '2':
                    courses[1]['len'] += 1
                elif str(c.group)[:1] == '3':
                    courses[2]['len'] += 1
                elif str(c.group)[:1] == '4':
                    courses[3]['len'] += 1
                elif str(c.group)[:1] == '5':
                    courses[4]['len'] += 1
                elif str(c.group)[:1] == '6':
                    courses[5]['len'] += 1

        hostel2 = [{'places': 0, 'persons': len(models.Person.query.filter_by(hostel_id=1).all())}]
        for room in models.Room.query.filter_by(hostel_id=1).all():
            hostel2[0]['places'] += int(room.numbers_of_person)

        hostel3 = [{'places': 0, 'persons': len(models.Person.query.filter_by(hostel_id=2).all())}]
        for room in models.Room.query.filter_by(hostel_id=2).all():
            hostel3[0]['places'] += int(room.numbers_of_person)

        hostel4 = [{'places': 0, 'persons': len(models.Person.query.filter_by(hostel_id=3).all())}]
        for room in models.Room.query.filter_by(hostel_id=3).all():
            hostel4[0]['places'] += int(room.numbers_of_person)

        foe = [{'foe': 'бюджет', 'len': len(models.Person.query.filter_by(form_of_education='б').all())},
               {'foe': 'контракт', 'len': len(models.Person.query.filter_by(form_of_education='к').all())}]

        windows_2 = [{'windows': len(models.Room.query.filter_by(windows=1, hostel_id=1).all()),
                      'rooms': len(models.Room.query.filter_by(hostel_id=1).all())}]
        windows_3 = [{'windows': len(models.Room.query.filter_by(windows=1, hostel_id=2).all()),
                      'rooms': len(models.Room.query.filter_by(hostel_id=2).all())}]
        windows_4 = [{'windows': len(models.Room.query.filter_by(windows=1, hostel_id=3).all()),
                      'rooms': len(models.Room.query.filter_by(hostel_id=3).all())}]

        hot_water_3 = [{'hot': len(models.Block.query.filter_by(hot_water=1, hostel_id=2).all()),
                        'blocks': len(models.Block.query.filter_by(hostel_id=2).all())}]
        hot_water_4 = [{'hot': len(models.Block.query.filter_by(hot_water=1, hostel_id=3).all()),
                        'blocks': len(models.Block.query.filter_by(hostel_id=3).all())}]

        econom_2 = [{'econom': len(models.Room.query.filter_by(econom=True, hostel_id=1).all()),
                     'rooms': len(models.Room.query.filter_by(hostel_id=1).all())}]
        econom_3 = [{'econom': len(models.Room.query.filter_by(econom=True, hostel_id=2).all()),
                     'rooms': len(models.Room.query.filter_by(hostel_id=2).all())}]
        econom_4 = [{'econom': len(models.Room.query.filter_by(econom=True, hostel_id=3).all()),
                     'rooms': len(models.Room.query.filter_by(hostel_id=3).all())}]

        return render_template('stat.html', stats=stats, departments=departments, rooms=rooms, courses=courses,
                               departments_2=departments_2, departments_3=departments_3, departments_4=departments_4,
                               hostel2=hostel2, hostel3=hostel3, hostel4=hostel4, foe=foe, windows_2=windows_2,
                               windows_3=windows_3, windows_4=windows_4, hot_water_3=hot_water_3,
                               hot_water_4=hot_water_4, econom_2=econom_2, econom_3=econom_3, econom_4=econom_4)


@app.route('/load')
@webLog
def load():
    from openpyxl import load_workbook
    import os
    from config import basedir
    wb = load_workbook(os.path.join(basedir, '1.xlsx'))
    ws = wb['БАЗА']
    for row in range(1, ws.max_row + 1):
        try:
            hostel_id = models.Hostel.query.filter_by(number=ws['G%s' % row].value).first().id
            room = models.Room.query.filter_by(
                hostel_id=hostel_id,
                room_number=ws['H%s' % row].value).first().id
        except:
            hostel_id = None
            print('ERROR', ws['A%s' % row].value)
            room = None
        m = models.Person(first_name=ws['A%s' % row].value,
                          last_name=ws['B%s' % row].value,
                          middle_name=ws['C%s' % row].value,
                          department=ws['D%s' % row].value,
                          group=ws['E%s' % row].value,
                          form_of_education=ws['F%s' % row].value,
                          hostel_id=hostel_id,
                          room_id=ws['H%s' % row].value,
                          birthday=ws['I%s' % row].value,
                          passport=ws['J%s' % row].value,
                          parents=ws['K%s' % row].value,
                          index=ws['L%s' % row].value,
                          region=ws['M%s' % row].value,
                          district=ws['N%s' % row].value,
                          settlement=ws['O%s' % row].value,
                          street=ws['P%s' % row].value,
                          phone_number_parent=ws['Q%s' % row].value,
                          phone_number=ws['R%s' % row].value,
                          note=ws['S%s' % row].value,
                          room=room)
        db.session.add(m)
        db.session.commit()


@app.route('/update')
@webLog
def update():
    m = models.Person.query.all()
    for item in m:
        item.hostel_id -= 1
        db.session.commit()


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1] in ALLOWED_EXTENSIONS


@app.route('/photo', methods=['GET', 'POST'])
@webLog
def upload_file():
    if request.method == 'POST':
        # check if the post request has the file part
        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        file = request.files['file']
        # if user does not select file, browser also
        # submit a empty part without filename
        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(
                str(datetime.utcnow()).replace(' ', '').replace('.', '_') + '.' + file.filename.rsplit('.', 1)[1])
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            db.session.add(models.Photo(path=filename, timestamp=datetime.utcnow()))
            db.session.commit()
            return redirect(url_for('upload_file'))
    else:
        photo = models.Photo.query.order_by(desc(models.Photo.timestamp)).limit(100).all()

        return render_template('photo.html', photos=photo, uf=app.config['UPLOAD_FOLDER'])


@app.route('/get_all_register', methods=['GET'])
@webLog
def get_all_register():
    family = models.Register_family.query.all()
    family_main = []
    student = models.Register_student.query.all()
    student_main = []
    if student:
        for stu in student:
            student_main.append([stu, models.Register_main.query.filter_by(id=stu.register_id).first()])

    if family:
        for fam in family:
            family_main.append([fam, models.Register_main.query.filter_by(id=fam.register_id).first()])

    return render_template('all_register.html', family=family_main, student=student_main)


@app.route('/profile/repair', methods=['GET', 'POST'])
@webLog
def repair():
    if request.method == 'GET':
        repairs = models.Repair.query.filter_by(person=current_user.person_id)
        return render_template('repair.html', current_user=current_user, repairs=repairs)
    elif request.method == 'POST':
        db.session.add(models.Repair(request.form['description'], request.form['tag'], current_user.person_id))
        db.session.commit()
        return redirect(url_for('repair'))


@app.route('/repair', methods=['GET', 'POST'])
@webLog
def all_repair():
    if request.method == "GET":
        repairs = db.session.query(models.Repair, models.Person, models.Room, models.Hostel).filter(
            models.Repair.fix == False).filter(models.Repair.person == models.Person.id).filter(
            models.Person.room == models.Room.id).filter(models.Room.hostel_id == models.Hostel.id).order_by(
            desc(models.Repair.open_date)).all()
        return render_template('all_repair.html', repairs=repairs)
    else:
        query = models.Repair.query.filter_by(id=request.form['id']).first()
        query.fix = True
        db.session.commit()
        return json.dumps({'status': 'OK'})


@app.route('/free', methods=['GET'])
@webLog
def free():
    query = db.session.query(models.Person, models.Room, models.Hostel) \
        .filter(models.Person.room_id == models.Room.id) \
        .filter(models.Hostel.id == models.Room.hostel_id) \
        .filter(models.Hostel.id == 3)
    print(query.all())


@app.route('/init')
@webLog
def init():
    for room in models.Room.query.all():
        places = room.numbers_of_person - len(models.Person.query.filter_by(room=room.id).all())
        db.session.add(models.Room_free(places=places, room_id=room.id))
        db.session.commit()


@app.route('/employees')
@webLog
def employees():
    return render_template('employees.html')


@app.route('/sliderpost/<id>')
@webLog
def sliderpost(id):
    post = models.News_Slider.query.filter_by(id=id).first()
    return render_template('post.html', post=post)


@app.route('/posts')
@app.route('/posts/page-<int:page>')
@webLog
def posts(page=1):
    PER_PAGE = 10
    posts = models.Post.query.order_by(desc(models.Post.timestamp)).paginate(page, PER_PAGE, False)
    return render_template('posts.html', posts=posts, page=page)


@app.route('/calendar_washing')
@login_required
@webLog
def calendar_washing():
    pass


@app.route('/new_wash', methods=['POST'])
def new_wash():
    from dateutil.parser import parse
    WASH_PER_DAY = 3
    if len(models.Washing.query.filter_by(start=parse(request.form['startDate'])).all()) < WASH_PER_DAY:
        db.session.add(models.Washing(start=parse(request.form['startDate']), end=parse(request.form['endDate']),
                                      person=request.form['name'], hostel=request.form['location'])
                       )
        db.session.commit()
        return json.dumps({'status': 'ok'})
    else:
        return abort(500)


@app.route('/new_get_person', methods=['POST'])
def new_get_person():
    query = db.session.query(models.Hostel, models.Person, models.Room).filter(
        models.Hostel.id == request.form['location']).filter(models.Person.id == request.form['id']).filter(
        models.Room.id == models.Person.room).first()

    return json.dumps({'status': 'ok',
                       'person': query[1].first_name + ' ' + query[1].last_name,
                       'room': query[2].room_number,
                       'hostel': query[0].number
                       })


@app.route('/api/news')
def api_news():
    from re import sub
    items = []
    for x in models.Post.query.order_by(desc(models.Post.timestamp)).all():
        items.append(
            {'id': x.id, 'title': x.title, 'previewtext': sub("<[^>]*>", '', x.previewtext),
             'body': sub("<[^>]*>", '', x.body).replace('\r\n', ''),
             'timestamp': x.timestamp, 'path': url_for('static', filename='files/' + x.path)})
    return jsonify({'status': 200, 'items': items, 'length': len(items)})


@app.route('/api/news/<int:id>')
def api_news_single(id):
    from re import sub
    x = models.Post.query.filter_by(id=id).first()
    return jsonify({'status': 200, 'item':
        {'id': x.id, 'title': x.title, 'previewtext': sub("<[^>]*>", '', x.previewtext),
         'body': sub("<[^>]*>", '', x.body).replace('\r\n', ''),
         'timestamp': x.timestamp, 'path': url_for('static', filename='files/' + x.path)}})
